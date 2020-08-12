#!/usr/bin/python3
# -*- coding: UTF-8 -*-

import openpyxl
import re
from info import Info


def load_excel():
    s = Info("Import excle")
    s.out_start()
    wb = openpyxl.load_workbook('/Users/kiana/Desktop/1.xlsx')
    sheet = wb.active
    s.out_end()
    return sheet


def get_ip_list(sheet):
    ip_list = []
    s = Info("Get IP list")
    s.out_start()
    for ip in sheet.iter_rows(2, sheet.max_row):
        ip_list.append(ip[0].value)
        # print(ip[0].value)
    s.out_end()
    return ip_list


def get_port_list(sheet):
    port_list = []
    pattern = re.compile('\d+')
    s = Info("Get Ports list")
    for port in sheet.iter_rows(2, sheet.max_row):
        str = port[3].value
        p = re.match(pattern, str)
        port_list.append(p.group(0))
        # print(p.group(0))
    s.out_end()
    return port_list


if __name__ == '__main__':
    sheet = load_excel()
    ip = get_ip_list(sheet)
    port = get_port_list(sheet)
    print(ip)
    print(port)

